"""
Module Exports - Exports Avancés Excel/PDF

Ce module fournit:
- Export Excel avec graphiques (openpyxl)
- Export PDF avec rapports professionnels (reportlab)
- Templates customisables
- Background jobs avec notifications
- Formats: CSV, Excel, PDF

Usage:
    from core.exports import ExportService, export_organisations_excel

    # Export Excel avec graphiques
    file_path = await export_organisations_excel(
        organisations=orgs,
        filename="organisations_2025.xlsx",
        include_charts=True
    )

    # Export PDF rapport
    pdf_path = await export_organisations_pdf(
        organisations=orgs,
        filename="rapport_organisations.pdf",
        template="standard"
    )
"""

import csv
import enum
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

# Excel
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors

# PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from models.mandat import Mandat
from models.organisation import Organisation
from models.person import Person


def _normalize_value(value: Any) -> Any:
    if isinstance(value, enum.Enum):
        return value.value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _row_from_item(item: Any) -> Dict[str, Any]:
    if isinstance(item, dict):
        source = item
    elif hasattr(item, "to_dict"):
        source = item.to_dict()
    else:
        source = {
            key: value
            for key, value in vars(item).items()
            if not key.startswith("_")
        }
    return {key: _normalize_value(value) for key, value in source.items()}


class ExportService:
    """
    Service d'export de données

    Gère les exports Excel, PDF et CSV avec mise en forme
    professionnelle et graphiques.
    """

    # ============================================
    # Export CSV
    # ============================================

    @staticmethod
    def export_csv(
        data: List[Any],
        filename: str = "export.csv",
        columns: Optional[List[str]] = None,
        headers: Optional[List[str]] = None,
    ) -> BytesIO:
        """
        Export CSV basique

        Args:
            data: Liste de dictionnaires ou d'objets SQLAlchemy
            columns: Colonnes à exporter (alias legacy)
            headers: Colonnes à exporter (ordre d'affichage)
            filename: Nom du fichier

        Returns:
            BytesIO: Contenu CSV
        """
        rows = [_row_from_item(item) for item in data]
        fieldnames = headers or columns
        if fieldnames is None:
            if rows:
                fieldnames = list(rows[0].keys())
            else:
                fieldnames = []

        text_buffer = StringIO()
        writer = csv.DictWriter(text_buffer, fieldnames=fieldnames, extrasaction="ignore")

        writer.writeheader()
        writer.writerows(rows)

        output = BytesIO()
        output.write(b'\xef\xbb\xbf')  # UTF-8 BOM pour Excel
        output.write(text_buffer.getvalue().encode("utf-8"))
        output.seek(0)
        return output

    # ============================================
    # Export Excel
    # ============================================

    @staticmethod
    def export_excel_simple(
        data: List[Any],
        filename: str,
        headers: Optional[List[str]] = None,
        sheet_name: str = "Données",
    ) -> BytesIO:
        """
        Export Excel simple sans graphiques

        Args:
            data: Données à exporter
            headers: Liste des colonnes à exporter (ordre)
            filename: Nom du fichier
            sheet_name: Nom de la feuille

        Returns:
            BytesIO: Fichier Excel
        """
        rows = [_row_from_item(item) for item in data]
        if headers is None:
            if rows:
                headers = list(rows[0].keys())
            else:
                headers = []

        display_headers = {field: field.replace("_", " ").title() for field in headers}

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style en-têtes
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Écrire en-têtes
        for col_idx, field in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, display_headers[field])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Écrire données
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, field in enumerate(headers, 1):
                value = row_data.get(field, '')
                ws.cell(row_idx, col_idx, value)

        # Auto-ajuster largeurs
        for col_idx, field in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20

        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def _get_enum_value(obj, attr_name: str, default="") -> str:
        """Extrait la valeur d'un enum ou retourne une string."""
        value = getattr(obj, attr_name, default)
        return value.value if isinstance(value, enum.Enum) else (value or default)

    @staticmethod
    def _extract_org_row_data(org) -> List[Any]:
        """Extrait les données d'une organisation pour une ligne Excel."""
        created_at = getattr(org, "created_at", None)
        return [
            getattr(org, "id", None),
            getattr(org, "name", ""),
            ExportService._get_enum_value(org, "type"),
            ExportService._get_enum_value(org, "category"),
            ExportService._get_enum_value(org, "pipeline_stage"),
            getattr(org, "email", "") or "",
            getattr(org, "phone", "") or "",
            getattr(org, "city", "") or "",
            getattr(org, "annual_revenue", ""),
            getattr(org, "employee_count", ""),
            created_at.strftime('%Y-%m-%d') if created_at else '',
            "Oui" if getattr(org, "is_active", False) else "Non"
        ]

    @staticmethod
    def _setup_excel_headers(ws, headers: List[str]) -> None:
        """Configure les en-têtes du fichier Excel."""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def export_organisations_excel(
        organisations: List[Organisation],
        filename: str = "organisations.xlsx",
        include_charts: bool = True,
    ) -> BytesIO:
        """
        Export Excel organisations avec graphiques

        Crée un fichier Excel avec:
        - Feuille "Organisations" (données)
        - Feuille "Statistiques" (graphiques)

        Args:
            organisations: Liste d'organisations
            filename: Nom du fichier
            include_charts: Inclure les graphiques

        Returns:
            BytesIO: Fichier Excel
        """
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Organisations"

        # En-têtes
        headers = [
            "ID", "Nom", "Type", "Catégorie", "Pipeline",
            "Email", "Téléphone", "Ville", "Revenus Annuels",
            "Nb Employés", "Date Création", "Actif"
        ]
        ExportService._setup_excel_headers(ws_data, headers)

        # Données
        for row_idx, org in enumerate(organisations, 2):
            row_data = ExportService._extract_org_row_data(org)
            for col_idx, value in enumerate(row_data, 1):
                ws_data.cell(row_idx, col_idx, value)

        # Auto-ajuster largeurs
        for col_idx in range(1, len(headers) + 1):
            ws_data.column_dimensions[get_column_letter(col_idx)].width = 18

        # === Feuille 2: Statistiques ===
        if include_charts:
            ws_stats = wb.create_sheet("Statistiques")

            # Compter par catégorie
            category_counts = {}
            for org in organisations:
                cat = getattr(org, 'category', 'Autre')
                if isinstance(cat, enum.Enum):
                    cat = cat.value
                category_counts[cat or 'Autre'] = category_counts.get(cat or 'Autre', 0) + 1

            # Écrire stats
            ws_stats.cell(1, 1, "Catégorie").font = Font(bold=True)
            ws_stats.cell(1, 2, "Nombre").font = Font(bold=True)

            for idx, (cat, count) in enumerate(category_counts.items(), 2):
                ws_stats.cell(idx, 1, cat)
                ws_stats.cell(idx, 2, count)

            # Graphique en barres
            chart = BarChart()
            chart.title = "Organisations par Catégorie"
            chart.x_axis.title = "Catégorie"
            chart.y_axis.title = "Nombre"

            data = Reference(ws_stats, min_col=2, min_row=1, max_row=len(category_counts) + 1)
            cats = Reference(ws_stats, min_col=1, min_row=2, max_row=len(category_counts) + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_stats.add_chart(chart, "D2")

        # Sauvegarder
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    # ============================================
    # Export PDF
    # ============================================

    @staticmethod
    def export_organisations_pdf(
        organisations: List[Organisation],
        filename: str = "organisations.pdf",
        title: str = "Rapport Organisations",
        author: Optional[str] = None,
    ) -> BytesIO:
        """
        Export PDF organisations avec rapport professionnel

        Args:
            organisations: Liste d'organisations
            filename: Nom du fichier
            title: Titre du rapport

        Returns:
            BytesIO: Fichier PDF
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        if author:
            doc.author = author
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=1,  # Center
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495E'),
            spaceAfter=12,
        )

        # Titre
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 20))

        # Résumé
        story.append(Paragraph("Résumé Exécutif", heading_style))
        story.append(Paragraph(
            f"Nombre total d'organisations: <b>{len(organisations)}</b>",
            styles['Normal']
        ))
        story.append(Spacer(1, 10))

        active_count = sum(1 for org in organisations if org.is_active)
        if organisations:
            active_ratio = active_count / len(organisations) * 100
        else:
            active_ratio = 0.0

        story.append(Paragraph(
            f"Organisations actives: <b>{active_count}</b> ({active_ratio:.1f}%)",
            styles['Normal']
        ))
        story.append(Spacer(1, 20))

        # Table des organisations
        story.append(Paragraph("Liste des Organisations", heading_style))

        # Préparer données table
        table_data = [['Nom', 'Type', 'Catégorie', 'Email', 'Actif']]

        for org in organisations[:50]:  # Limiter à 50 pour PDF
            org_type = getattr(org, 'type', '')
            if isinstance(org_type, enum.Enum):
                org_type = org_type.value
            category = getattr(org, 'category', '')
            if isinstance(category, enum.Enum):
                category = category.value
            email = getattr(org, 'email', '') or ''

            table_data.append([
                (getattr(org, 'name', '') or '')[:30],
                org_type,
                category,
                email[:25],
                '✓' if getattr(org, 'is_active', False) else '✗',
            ])

        # Créer table
        table = Table(table_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.8*inch, 0.5*inch])
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

            # Corps
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),

            # Bordures
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECF0F1')]),
        ]))

        story.append(table)

        if len(organisations) > 50:
            story.append(Spacer(1, 10))
            story.append(Paragraph(
                f"<i>Note: Seules les 50 premières organisations sont affichées. Total: {len(organisations)}</i>",
                styles['Normal']
            ))

        # Générer PDF
        doc.build(story)
        output.seek(0)

        return output

    @staticmethod
    def export_mandats_pdf(
        mandats: List[Mandat],
        filename: str = "mandats.pdf",
    ) -> BytesIO:
        """
        Export PDF mandats

        Args:
            mandats: Liste de mandats
            filename: Nom du fichier

        Returns:
            BytesIO: Fichier PDF
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []

        styles = getSampleStyleSheet()

        # Titre
        story.append(Paragraph("Rapport Mandats", styles['Title']))
        story.append(Spacer(1, 20))

        # Table
        table_data = [['Numéro', 'Type', 'Status', 'Organisation', 'Montant']]

        for mandat in mandats[:50]:
            table_data.append([
                mandat.number if hasattr(mandat, 'number') else '',
                mandat.type if hasattr(mandat, 'type') else '',
                mandat.status if hasattr(mandat, 'status') else '',
                mandat.organisation.name if hasattr(mandat, 'organisation') else '',
                f"{mandat.amount} €" if hasattr(mandat, 'amount') else '',
            ])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        story.append(table)

        doc.build(story)
        output.seek(0)

        return output


# ============================================
# Helpers Simplifiés
# ============================================

async def export_organisations_csv(
    organisations: List[Organisation],
    db: Session | None = None,
    filename: str = "organisations.csv",
) -> BytesIO:
    """Helper CSV pour les organisations."""

    columns = [
        "id",
        "name",
        "type",
        "category",
        "email",
        "city",
        "country",
        "is_active",
    ]

    data = [org.to_dict() for org in organisations]

    return ExportService.export_csv(data=data, columns=columns, filename=filename)


async def export_organisations_excel(
    organisations: List[Organisation],
    db: Session | None = None,
    filename: str = "organisations.xlsx",
    include_charts: bool = True,
) -> BytesIO:
    """
    Helper export Excel organisations

    Args:
        organisations: Liste d'organisations
        filename: Nom du fichier
        include_charts: Inclure graphiques

    Returns:
        BytesIO: Fichier Excel
    """
    return ExportService.export_organisations_excel(
        organisations=organisations,
        filename=filename,
        include_charts=include_charts,
    )


async def export_organisations_pdf(
    organisations: List[Organisation],
    db: Session | None = None,
    filename: str = "organisations.pdf",
    title: str = "Rapport Organisations",
) -> BytesIO:
    """
    Helper export PDF organisations

    Args:
        organisations: Liste d'organisations
        filename: Nom du fichier
        title: Titre du rapport

    Returns:
        BytesIO: Fichier PDF
    """
    return ExportService.export_organisations_pdf(
        organisations=organisations,
        filename=filename,
        title=title,
    )


async def export_mandats_pdf(
    mandats: List[Mandat],
    db: Session | None = None,
    filename: str = "mandats.pdf",
) -> BytesIO:
    """
    Helper export PDF mandats.
    """
    return ExportService.export_mandats_pdf(
        mandats=mandats,
        filename=filename,
    )
